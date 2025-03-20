from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from dataclasses import dataclass
from typing import Dict, List, Optional
import time
import logging
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

@dataclass
class Product:
    name: str
    price: float
    currency: str
    packaging: str
    min_order: str
    url: str
    category: str
    subcategory: str

class JettaParser:
    def __init__(self):
        self.service = Service(ChromeDriverManager().install())
        self.driver = None
        self.wait = None
        self.excel_file = f"jetta_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.current_row = 2  # Глобальный счетчик строк, начинаем с 2 (строка 1 - заголовки)
        self.setup_logging()
        self.setup_excel()

    def setup_logging(self):
        """Setup logging to both file and console."""
        log_file = f"jetta_parser_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        
        # Create formatters
        file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        console_formatter = logging.Formatter('%(message)s')
        
        # Setup file handler
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(file_formatter)
        
        # Setup console handler
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(console_formatter)
        
        # Setup logger
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        # Log the Excel filename
        self.logger.info(f"Excel file will be saved as: {self.excel_file}")

    def setup_driver(self):
        """Setup Chrome driver in headless mode."""
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')  # Run in headless mode
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        
        self.driver = webdriver.Chrome(service=self.service, options=chrome_options)
        self.wait = WebDriverWait(self.driver, 10)
        self.logger.info("Chrome driver initialized in headless mode")

    def setup_excel(self):
        """Initialize Excel file with headers."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Add headers
            headers = ["Категория", "Подкатегория", "Название", "Цена", "Валюта", "Фасовка", "Минимальный заказ", "URL"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            wb.save(self.excel_file)
            self.logger.info(f"Excel file created: {self.excel_file}")
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {e}")

    def save_product_to_excel(self, product: Product, row: int):
        """Save single product data to Excel file."""
        try:
            # Check if file exists
            if not os.path.exists(self.excel_file):
                self.setup_excel()
            
            # Load existing workbook
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Add data
            ws.cell(row=row, column=1, value=product.category)
            ws.cell(row=row, column=2, value=product.subcategory)
            ws.cell(row=row, column=3, value=product.name)
            ws.cell(row=row, column=4, value=product.price)
            ws.cell(row=row, column=5, value=product.currency)
            ws.cell(row=row, column=6, value=product.packaging)
            ws.cell(row=row, column=7, value=product.min_order)
            ws.cell(row=row, column=8, value=product.url)
            
            # Save and close the workbook
            wb.save(self.excel_file)
            wb.close()
            self.logger.info(f"Product saved to Excel: {product.name}")
        except Exception as e:
            self.logger.error(f"Error saving product to Excel: {e}")
            # Try to close the workbook if it's open
            try:
                wb.close()
            except:
                pass

    def get_category_links(self) -> Dict[str, str]:
        """Get all category links from the main page."""
        try:
            self.driver.get('https://jetta-chemical.ru/katalog')
            elements = self.wait.until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "content_main_ngroups_title"))
            )
            
            links_dict = {}
            for element in elements:
                title = element.text
                link = element.find_element(By.TAG_NAME, 'a').get_attribute('href')
                links_dict[title] = link
                
            self.logger.info(f"Found {len(links_dict)} categories")
            return links_dict
        except Exception as e:
            self.logger.error(f"Error getting category links: {e}")
            return {}

    def get_subcategory_links(self, category_url: str) -> Dict[str, str]:
        """Get all subcategory links from a category page."""
        try:
            self.driver.get(category_url)
            time.sleep(2)  # Wait for dynamic content to load
            
            elements = self.wait.until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "catalog_group_list_item"))
            )
            
            links_dict = {}
            for element in elements:
                title = element.text
                link = element.find_element(By.TAG_NAME, 'a').get_attribute('href')
                links_dict[title] = link
                
            self.logger.info(f"Found {len(links_dict)} subcategories")
            return links_dict
        except Exception as e:
            self.logger.error(f"Error getting subcategory links: {e}")
            return {}

    def parse_product(self, product_url: str, category: str, subcategory: str) -> Optional[Product]:
        """Parse individual product data from its URL."""
        try:
            # Get product name
            name = self.driver.find_element(By.CLASS_NAME, "content_navi_title").text
            url = product_url
            
            # Get price and currency
            price_element = self.driver.find_element(By.CLASS_NAME, "item_page_cd1_d2_i1_d1price_d2")
            price_text = price_element.text.replace(' ', '').replace(',', '.').replace('₽', '').strip()
            price = float(price_text)
            currency = "RUB"
            
            # Get packaging and min order
            info_text = self.driver.find_element(By.CLASS_NAME, "item-page_smalldescription").text
            packaging = "Не указано"
            min_order = "Не указано"
            
            for line in info_text.split('\n'):
                if 'Фасовка:' in line:
                    packaging = line.split(':')[1].strip()
                if 'Минимальный заказ:' in line:
                    min_order = line.split(':')[1].strip()
            
            return Product(
                name=name,
                price=price,
                currency=currency,
                packaging=packaging,
                min_order=min_order,
                url=url,
                category=category,
                subcategory=subcategory
            )
        except Exception as e:
            self.logger.error(f"Error parsing product: {e}")
            return None

    def get_products(self, subcategory_url: str, category: str, subcategory: str) -> List[Product]:
        """Get all products from a subcategory page."""
        try:
            self.driver.get(subcategory_url)
            time.sleep(2)  # Wait for dynamic content to load
            
            # Click "Load more" button until all products are loaded
            while True:
                try:
                    load_more_button = self.driver.find_element(By.CLASS_NAME, "catalog_more_bott.more_bott_off")
                    self.driver.execute_script("arguments[0].click();", load_more_button)
                    time.sleep(2)  # Wait for new products to load
                except:
                    break  # Button not found or not clickable anymore
                    
            # Get all product elements after loading everything
            product_elements = self.wait.until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "catalog_item_v_block_d4_cd_d1"))
            )

            product_elements_links = [element.find_element(By.TAG_NAME, 'a').get_attribute('href') for element in product_elements]
            
            products = []
            for element in product_elements_links:
                try:
                    self.driver.get(element)
                    time.sleep(2)  # Reduced wait time
                    product = self.parse_product(element, category, subcategory)
                    if product:
                        products.append(product)
                        self.save_product_to_excel(product, self.current_row)
                        self.logger.info(f"Processed product {self.current_row}: {product.name}")
                        self.logger.info(f"Verification - Product {product.name} saved to row {self.current_row}")
                        self.current_row += 1  # Увеличиваем счетчик строк
                except Exception as e:
                    self.logger.error(f"Error processing product element: {e}")
                    continue
                        
            self.logger.info(f"Found {len(products)} products")
            return products
        except Exception as e:
            self.logger.error(f"Error getting products: {e}")
            return []

    def parse_all(self):
        """Main method to parse all products from the website."""
        try:
            self.setup_driver()
            
            # Get all categories
            categories = self.get_category_links()
            
            all_products = []
            for category_name, category_url in categories.items():
                self.logger.info(f"Processing category: {category_name}")
                
                # Get all subcategories
                subcategories = self.get_subcategory_links(category_url)
                
                for subcategory_name, subcategory_url in subcategories.items():
                    self.logger.info(f"Processing subcategory: {subcategory_name}")
                    
                    # Get all products
                    products = self.get_products(subcategory_url, category_name, subcategory_name)
                    all_products.extend(products)
                    
            self.logger.info(f"Total products parsed: {len(all_products)}")
            return all_products
            
        except Exception as e:
            self.logger.error(f"Error during parsing: {e}")
            return []
            
        finally:
            if self.driver:
                self.driver.quit()

if __name__ == "__main__":
    parser = JettaParser()
    products = parser.parse_all()
    
    # Print results
    # for product in products:
    #     print(f"\nProduct: {product.name}")
    #     print(f"Price: {product.price} {product.currency}")
    #     print(f"Packaging: {product.packaging}")
    #     print(f"Min Order: {product.min_order}")
    #     print(f"URL: {product.url}")